from __future__ import annotations

import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def dedupe_results(rows: list[dict]) -> list[dict]:
    """
    Прибирає дублікати.
    - Для detail_view: унікально по part_no
    - Для modal: унікально по (part_no, Location)
    Залишає останній запис (last wins).
    """
    seen = {}
    for r in rows:
        part_no = str(r.get("part_no") or "").strip()
        loc = str(r.get("Location") or "").strip()

        if loc:
            key = (part_no, loc)  # modal rows
        else:
            key = (part_no, None)  # detail row

        # last wins
        seen[key] = r

    return list(seen.values())


def save_costex_results_xlsx(
    rows: list[dict[str, Any]],
    out_dir: str | Path = ".",
    archive_dir_name: str = "archive",
    latest_name: str = "costex_catalog_latest.xlsx",
) -> tuple[Path, Path]:
    """
    Записує результати у два XLSX файли з ФІКСОВАНОЮ структурою колонок:
    Reference | Product Name | Retail Price Tax Exc | Quantity | Images | Features |
    Width | Height | Depth | Weight | Default Category | Categories
    """

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    archive_dir = out_dir / archive_dir_name
    archive_dir.mkdir(parents=True, exist_ok=True)

    today = datetime.now().strftime("%Y%m%d")
    dated_name = f"costex_catalog_{today}.xlsx"

    latest_path = out_dir / latest_name
    dated_path = out_dir / dated_name

    # ---------- ARCHIVE OLD FILES ----------
    pat = re.compile(r"^costex_catalog_.*\.xlsx$", re.IGNORECASE)
    keep = {latest_path.name.lower(), dated_path.name.lower()}

    for p in out_dir.iterdir():
        if not p.is_file():
            continue
        if not pat.match(p.name):
            continue
        if p.name.lower() in keep:
            continue

        dest = archive_dir / p.name
        if dest.exists():
            ts = datetime.now().strftime("%H%M%S")
            dest = archive_dir / f"{p.stem}_{ts}{p.suffix}"

        shutil.move(str(p), str(dest))

    # ---------- FIXED STRUCTURE ----------
    headers = [
        "Reference",
        "Product Name",
        "Retail Price Tax Exc",
        "Quantity",
        "Images",
        "Features",
        "Width",
        "Height",
        "Depth",
        "Weight",
        "Default Category",
        "Categories",
    ]

    field_map = {
        "Reference": "part_no",
        "Product Name": "Title",
        "Retail Price Tax Exc": "Unit Price",
        "Quantity": "Qty Available",
        "Images": "Image URL",
        "Features": "Features",
        "Width": "Width",
        "Height": "Height",
        "Depth": "Depth",
        "Weight": "Lbs",
        "Default Category": "Category",
        "Categories": "Categories",
    }

    def _write_xlsx(path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "catalog"

        # header
        ws.append(headers)

        # rows
        for r in rows:
            ws.append([
                r.get(field_map[h], "") for h in headers
            ])

        # auto width
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)
            for cell in ws[get_column_letter(col_idx)]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        wb.save(path)

    _write_xlsx(latest_path)
    _write_xlsx(dated_path)

    return latest_path, dated_path
